# main.py
import json
import os
import pandas as pd
import threading
import queue
import time
from concurrent.futures import ThreadPoolExecutor
from axies_API import AxesExtractor
from curve_API import CurveAnalysisAPI

class PipelineProcessor:
    """
    流水线式多线程处理器
    """
    def __init__(self, max_workers=4):
        self.max_workers = max_workers
        self.queues = {
            'step1': queue.Queue(),  # 步骤1队列：坐标轴提取
            'step2': queue.Queue(),  # 步骤2队列：曲线提取
            'step3': queue.Queue(),  # 步骤3队列：坐标转换
            'step4': queue.Queue(),  # 步骤4队列：数据构建
            'completed': queue.Queue()  # 完成队列
        }
        self.locks = {
            'step1': threading.Lock(),
            'step2': threading.Lock(),
            'step3': threading.Lock(),
            'step4': threading.Lock()
        }
        self.results = {}
        self.order_tracker = {}
        self.next_expected_index = 0
        self.completed_count = 0
        self.total_images = 0
        self.running = False

    def safe_format(self, value, format_spec=".2f"):
        """
        安全格式化函数，处理None值
        """
        if value is None:
            return "None"
        try:
            return format(value, format_spec)
        except (TypeError, ValueError):
            return str(value)

    def step1_extract_axes(self, image_info):
        """
        步骤1：提取坐标轴信息
        """
        image_index, image_path = image_info
        print(f"步骤1: 正在处理图像 {image_index+1}/{self.total_images} - {os.path.basename(image_path)}")
        
        # 验证图像文件存在
        if not os.path.exists(image_path):
            error_msg = f"图像文件不存在: {image_path}"
            print(f"错误: {error_msg}")
            return {'success': False, 'error': error_msg, 'image_index': image_index}
        
        try:
            axes_extractor = AxesExtractor(debug=False)
            axes_info = axes_extractor.extract_axes_info(image_path)
            
            if not axes_info['success']:
                error_msg = f"坐标轴提取失败: {axes_info['message']}"
                print(f"错误: {error_msg}")
                return {'success': False, 'error': error_msg, 'image_index': image_index}
            
            # 获取格式化后的坐标轴参数
            axes_params = axes_extractor.get_formatted_results(axes_info)
            
            # 提取关键参数，使用安全默认值
            x_left = axes_params['x_axis']['limits']['left'] if axes_params['x_axis']['limits']['left'] is not None else 0
            x_right = axes_params['x_axis']['limits']['right'] if axes_params['x_axis']['limits']['right'] is not None else 0
            x_title = axes_params['x_axis']['title'] if axes_params['x_axis']['title'] is not None else "未识别"
            x_pixels_per_value = axes_params['x_axis']['scale']['pixels_per_value'] if axes_params['x_axis']['scale']['pixels_per_value'] is not None else 0
            
            y_bottom = axes_params['y_axis']['limits']['bottom'] if axes_params['y_axis']['limits']['bottom'] is not None else 0
            y_top = axes_params['y_axis']['limits']['top'] if axes_params['y_axis']['limits']['top'] is not None else 0
            y_title = axes_params['y_axis']['title'] if axes_params['y_axis']['title'] is not None else "未识别"
            y_pixels_per_value = axes_params['y_axis']['scale']['pixels_per_value'] if axes_params['y_axis']['scale']['pixels_per_value'] is not None else 0
            
            print(f"步骤1完成: 图像 {image_index+1} - 坐标轴信息提取成功")
            
            return {
                'success': True,
                'image_index': image_index,
                'image_path': image_path,
                'axes_params': axes_params,
                'x_left': x_left,
                'x_right': x_right,
                'x_title': x_title,
                'x_pixels_per_value': x_pixels_per_value,
                'y_bottom': y_bottom,
                'y_top': y_top,
                'y_title': y_title,
                'y_pixels_per_value': y_pixels_per_value
            }
            
        except Exception as e:
            error_msg = f"步骤1处理异常: {e}"
            print(f"错误: {error_msg}")
            return {'success': False, 'error': error_msg, 'image_index': image_index}

    def step2_extract_curves(self, step1_result):
        """
        步骤2：提取曲线数据
        """
        if not step1_result['success']:
            return step1_result  # 直接传递错误结果
            
        image_index = step1_result['image_index']
        image_path = step1_result['image_path']
        
        print(f"步骤2: 正在处理图像 {image_index+1}/{self.total_images} - {os.path.basename(image_path)}")
        
        try:
            curve_api = CurveAnalysisAPI()
            curve_results = curve_api.analyze_image(
                image_path=image_path,
                yolo_threshold=0.3,
                curve_points=128,
                color_tolerance=40,
                visualize=False
            )
            
            if 'error' in curve_results:
                error_msg = f"曲线提取失败: {curve_results['error']}"
                print(f"错误: {error_msg}")
                return {'success': False, 'error': error_msg, 'image_index': image_index}
            
            print(f"步骤2完成: 图像 {image_index+1} - 找到 {curve_results['total_legends']} 条曲线")
            
            # 合并步骤1和步骤2的结果
            result = step1_result.copy()
            result['curve_results'] = curve_results
            return result
            
        except Exception as e:
            error_msg = f"步骤2处理异常: {e}"
            print(f"错误: {error_msg}")
            return {'success': False, 'error': error_msg, 'image_index': image_index}

    def step3_coordinate_transform(self, step2_result):
        """
        步骤3：坐标转换
        """
        if not step2_result['success']:
            return step2_result  # 直接传递错误结果
            
        image_index = step2_result['image_index']
        image_path = step2_result['image_path']
        
        print(f"步骤3: 正在处理图像 {image_index+1}/{self.total_images} - {os.path.basename(image_path)}")
        
        try:
            # 从步骤1结果中获取坐标轴参数
            x_left = step2_result['x_left']
            x_right = step2_result['x_right']
            x_pixels_per_value = step2_result['x_pixels_per_value']
            y_bottom = step2_result['y_bottom']
            y_top = step2_result['y_top']
            y_pixels_per_value = step2_result['y_pixels_per_value']
            
            # 从步骤2结果中获取曲线数据
            curve_results = step2_result['curve_results']
            
            # 坐标转换
            transformed_curves = []
            
            for curve in curve_results['curves']:
                legend_name = curve['legend_name']
                original_points = curve['points']
                
                transformed_points = []
                for point in original_points:
                    try:
                        # 严格按照指定的公式进行坐标转换
                        new_x = (point[0] - x_left) * x_pixels_per_value
                        new_y = (point[1] - y_bottom) * y_pixels_per_value
                        transformed_points.append([float(new_x), float(new_y)])
                    except Exception as e:
                        print(f"    坐标转换错误: {e}, 跳过该点")
                        continue
                
                if transformed_points:
                    transformed_curve = {
                        'legend_name': legend_name,
                        'original_points_count': len(original_points),
                        'transformed_points': transformed_points
                    }
                    transformed_curves.append(transformed_curve)
                else:
                    print(f"    警告: 图例 '{legend_name}' 没有成功转换的点")
            
            if not transformed_curves:
                error_msg = "没有成功转换任何曲线数据"
                print(f"错误: {error_msg}")
                return {'success': False, 'error': error_msg, 'image_index': image_index}
            
            print(f"步骤3完成: 图像 {image_index+1} - 坐标转换完成")
            
            # 合并结果
            result = step2_result.copy()
            result['transformed_curves'] = transformed_curves
            return result
            
        except Exception as e:
            error_msg = f"步骤3处理异常: {e}"
            print(f"错误: {error_msg}")
            return {'success': False, 'error': error_msg, 'image_index': image_index}

    def step4_build_excel_data(self, step3_result):
        """
        步骤4：构建Excel数据
        """
        if not step3_result['success']:
            return step3_result  # 直接传递错误结果
            
        image_index = step3_result['image_index']
        image_path = step3_result['image_path']
        
        print(f"步骤4: 正在处理图像 {image_index+1}/{self.total_images} - {os.path.basename(image_path)}")
        
        try:
            # 从之前的结果中获取必要信息
            x_title = step3_result['x_title']
            y_title = step3_result['y_title']
            transformed_curves = step3_result['transformed_curves']
            
            # 构建Excel数据
            excel_data = []
            image_name = os.path.basename(image_path)
            image_id = os.path.splitext(image_name)[0]
            
            for curve in transformed_curves:
                points_str = json.dumps(curve['transformed_points'], ensure_ascii=False)
                
                row_data = {
                    'figure_name (id)': f"{image_name} ({image_id})",
                    'figure_index': '',  # 空着不填
                    'figure_title': '',  # 空着不填
                    'x-label': x_title,
                    'y-label': y_title,
                    'sample': curve['legend_name'],
                    'point_coordinates': points_str,
                    'note': ''  # 空着不填
                }
                
                excel_data.append(row_data)
            
            print(f"步骤4完成: 图像 {image_index+1} - 生成 {len(excel_data)} 行数据")
            
            # 最终结果
            final_result = {
                'success': True,
                'image_index': image_index,
                'image_path': image_path,
                'excel_data': excel_data,
                'image_name': image_name,
                'x_title': x_title,
                'y_title': y_title,
                'curves_count': len(transformed_curves),
                'image_id': image_id
            }
            
            return final_result
            
        except Exception as e:
            error_msg = f"步骤4处理异常: {e}"
            print(f"错误: {error_msg}")
            return {'success': False, 'error': error_msg, 'image_index': image_index}

    def worker_step1(self):
        """步骤1工作线程"""
        while self.running:
            try:
                image_info = self.queues['step1'].get(timeout=1)
                if image_info is None:  # 结束信号
                    break
                    
                result = self.step1_extract_axes(image_info)
                
                # 按顺序放入步骤2队列
                with self.locks['step2']:
                    self.queues['step2'].put(result)
                    
            except queue.Empty:
                continue

    def worker_step2(self):
        """步骤2工作线程"""
        while self.running:
            try:
                step1_result = self.queues['step2'].get(timeout=1)
                if step1_result is None:  # 结束信号
                    break
                    
                result = self.step2_extract_curves(step1_result)
                
                # 按顺序放入步骤3队列
                with self.locks['step3']:
                    self.queues['step3'].put(result)
                    
            except queue.Empty:
                continue

    def worker_step3(self):
        """步骤3工作线程"""
        while self.running:
            try:
                step2_result = self.queues['step3'].get(timeout=1)
                if step2_result is None:  # 结束信号
                    break
                    
                result = self.step3_coordinate_transform(step2_result)
                
                # 按顺序放入步骤4队列
                with self.locks['step4']:
                    self.queues['step4'].put(result)
                    
            except queue.Empty:
                continue

    def worker_step4(self):
        """步骤4工作线程"""
        while self.running:
            try:
                step3_result = self.queues['step4'].get(timeout=1)
                if step3_result is None:  # 结束信号
                    break
                    
                result = self.step4_build_excel_data(step3_result)
                
                # 按顺序处理完成结果
                self.process_completed_result(result)
                
            except queue.Empty:
                continue

    def process_completed_result(self, result):
        """处理完成的结果，确保按顺序输出"""
        image_index = result['image_index']
        
        # 等待轮到当前图像
        while self.next_expected_index != image_index and self.running:
            time.sleep(0.1)  # 短暂等待
        
        # 存储结果
        self.results[image_index] = result
        self.next_expected_index += 1
        self.completed_count += 1
        
        # 放入完成队列
        self.queues['completed'].put(result)
        
        status = "成功" if result['success'] else "失败"
        print(f"✓ 图像 {image_index+1}/{self.total_images} 处理完成 - 状态: {status}")

    def process_all_images(self, figs_folder="figs", output_file="results.xlsx"):
        """
        使用流水线多线程处理所有图像
        """
        # 检查figs文件夹是否存在
        if not os.path.exists(figs_folder):
            print(f"错误: 文件夹 '{figs_folder}' 不存在")
            return
        
        # 支持的图片格式
        supported_formats = {'.jpg', '.jpeg', '.tif', '.tiff', '.png'}
        
        # 获取所有图片文件
        image_files = []
        for file in os.listdir(figs_folder):
            file_ext = os.path.splitext(file)[1].lower()
            if file_ext in supported_formats:
                image_files.append(os.path.join(figs_folder, file))
        
        # 按文件名排序
        image_files.sort()
        
        if not image_files:
            print(f"在文件夹 '{figs_folder}' 中没有找到支持的图片文件")
            print(f"支持的格式: {', '.join(supported_formats)}")
            return
        
        self.total_images = len(image_files)
        self.running = True
        
        print(f"找到 {self.total_images} 个图片文件，开始流水线处理...")
        for i, img_file in enumerate(image_files):
            print(f"  {i+1:2d}. {os.path.basename(img_file)}")
        
        # 启动工作线程
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 启动每个步骤的工作线程
            step1_future = executor.submit(self.worker_step1)
            step2_future = executor.submit(self.worker_step2)
            step3_future = executor.submit(self.worker_step3)
            step4_future = executor.submit(self.worker_step4)
            
            # 将图像任务放入步骤1队列
            print("\n开始流水线处理...")
            start_time = time.time()
            
            for i, image_path in enumerate(image_files):
                self.queues['step1'].put((i, image_path))
            
            # 等待所有任务完成
            while self.completed_count < self.total_images and self.running:
                time.sleep(0.5)
                elapsed = time.time() - start_time
                print(f"\r处理进度: {self.completed_count}/{self.total_images} [已用时间: {elapsed:.1f}s]", end="", flush=True)
            
            # 发送结束信号
            self.queues['step1'].put(None)
            self.queues['step2'].put(None)
            self.queues['step3'].put(None)
            self.queues['step4'].put(None)
            
            self.running = False
            
            # 等待所有线程完成
            step1_future.result()
            step2_future.result()
            step3_future.result()
            step4_future.result()
        
        total_time = time.time() - start_time
        print(f"\n\n所有图像处理完成! 总用时: {total_time:.2f}秒")
        
        # 收集处理结果
        processed_results = []
        failed_images = []
        all_excel_data = []
        
        for i in range(self.total_images):
            if i in self.results:
                result = self.results[i]
                if result['success']:
                    processed_results.append(result)
                    all_excel_data.extend(result['excel_data'])
                else:
                    failed_images.append(os.path.basename(result['image_path']))
        
        # 保存结果到Excel
        if all_excel_data:
            self.save_to_excel(all_excel_data, processed_results, failed_images, output_file)
            
            # 显示处理统计信息
            self.print_processing_summary(processed_results, failed_images, total_time)
            
            # 可选：打印详细坐标信息
            self.print_detailed_coordinates(all_excel_data, max_curves=3, max_points=5)
        
        return {
            'all_excel_data': all_excel_data,
            'processed_results': processed_results,
            'failed_images': failed_images,
            'total_images': len(processed_results),
            'total_curves': len(all_excel_data)
        }

    def save_to_excel(self, all_excel_data, processed_results, failed_images, output_file):
        """保存结果到Excel文件"""
        print(f"\n保存所有结果到Excel文件: {output_file}")
        
        try:
            df = pd.DataFrame(all_excel_data)
            
            # 设置Excel写入器，优化格式
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='All_Curve_Data', index=0)
                
                # 自动调整列宽
                worksheet = writer.sheets['All_Curve_Data']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✓ Excel文件保存成功: {output_file}")
            
        except Exception as e:
            print(f"✗ 保存Excel文件失败: {e}")
            return
        
        # 创建汇总工作表
        self.create_summary_sheet(processed_results, output_file, failed_images)

    def print_processing_summary(self, processed_results, failed_images, total_time):
        """打印处理汇总信息"""
        print("\n" + "=" * 80)
        print("处理汇总信息")
        print("=" * 80)
        
        print(f"成功处理的图片数量: {len(processed_results)}")
        print(f"处理失败的图片数量: {len(failed_images)}")
        print(f"总处理时间: {total_time:.2f}秒")
        
        if failed_images:
            print("失败的图片:")
            for img in failed_images:
                print(f"  - {img}")
        
        total_curves = sum(result['curves_count'] for result in processed_results)
        print(f"总曲线数量: {total_curves}")
        
        if processed_results:
            print("\n各图片处理详情:")
            print("-" * 100)
            print(f"{'图片名称':<30} {'X轴标题':<20} {'Y轴标题':<20} {'曲线数量':<10}")
            print("-" * 100)
            
            for result in processed_results:
                print(f"{result['image_name']:<30} {result['x_title']:<20} {result['y_title']:<20} {result['curves_count']:<10}")

    def print_detailed_coordinates(self, all_excel_data, max_curves=3, max_points=5):
        """打印详细的坐标点信息"""
        print("\n" + "=" * 60)
        print("详细坐标点信息（预览）")
        print("=" * 60)
        
        for i, row in enumerate(all_excel_data[:max_curves]):
            try:
                points = json.loads(row['point_coordinates'])
                print(f"\n曲线 {i+1}: {row['sample']} (来自 {row['figure_name (id)']})")
                print(f"坐标点 (显示前{min(max_points, len(points))}个):")
                
                for j, point in enumerate(points[:max_points]):
                    print(f"  点 {j+1:3d}: X = {self.safe_format(point[0], '12.6f')}, Y = {self.safe_format(point[1], '12.6f')}")
                
                if len(points) > max_points:
                    print(f"  ... (共{len(points)}个点)")
            except Exception as e:
                print(f"\n曲线 {i+1} 坐标解析错误: {e}")
        
        if len(all_excel_data) > max_curves:
            print(f"\n... (共{len(all_excel_data)}条曲线，完整数据请查看Excel文件)")

    def create_summary_sheet(self, processed_results, output_filename, failed_images):
        """创建汇总信息工作表"""
        try:
            from openpyxl import load_workbook
            
            # 加载现有的Excel文件
            wb = load_workbook(output_filename)
            
            # 创建汇总工作表
            ws_summary = wb.create_sheet(title="Summary", index=1)
            
            # 添加汇总信息
            summary_data = [
                ["图像分析结果汇总"],
                [""],
                ["处理时间:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["处理模式:", "流水线多线程"],
                ["成功处理的图片数量:", len(processed_results)],
                ["处理失败的图片数量:", len(failed_images)],
                ["总曲线数量:", sum(result['curves_count'] for result in processed_results)],
                [""],
                ["成功处理的图片列表:"],
                ["图片名称", "X轴标题", "Y轴标题", "曲线数量"]
            ]
            
            # 添加每个图片的信息
            for result in processed_results:
                summary_data.append([
                    result['image_name'],
                    result['x_title'],
                    result['y_title'],
                    result['curves_count']
                ])
            
            if failed_images:
                summary_data.extend([
                    [""],
                    ["处理失败的图片列表:"]
                ])
                for img in failed_images:
                    summary_data.append([img])
            
            summary_data.extend([
                [""],
                ["数据说明:"],
                ["- figure_name (id): 图像名称和ID"],
                ["- figure_index: 图像索引（暂空）"],
                ["- figure_title: 图像标题（暂空）"],
                ["- x-label: X轴标题"],
                ["- y-label: Y轴标题"],
                ["- sample: 图例/样本名称"],
                ["- point_coordinates: 转换后的坐标点集合"],
                ["- note: 备注（暂空）"]
            ])
            
            # 写入汇总数据
            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws_summary.cell(row=row_idx, column=col_idx, value=value)
            
            # 设置标题样式
            from openpyxl.styles import Font
            ws_summary['A1'].font = Font(bold=True, size=14)
            
            # 自动调整列宽
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # 保存修改
            wb.save(output_filename)
            print(f"✓ 已添加汇总工作表到: {output_filename}")
            
        except ImportError:
            print("警告: 未安装openpyxl，无法添加汇总工作表")
        except Exception as e:
            print(f"添加汇总工作表时出错: {e}")

def process_all_images(figs_folder="figs", output_file="results.xlsx", max_workers=4):
    """
    使用流水线多线程处理所有图像的主函数
    """
    processor = PipelineProcessor(max_workers=max_workers)
    return processor.process_all_images(figs_folder, output_file)

if __name__ == "__main__":
    try:
        # 使用流水线多线程处理figs文件夹中的所有图片
        result = process_all_images(figs_folder="figs", output_file="results.xlsx", max_workers=4)
        
    except Exception as e:
        print(f"程序执行过程中出错: {e}")
        import traceback
        traceback.print_exc()