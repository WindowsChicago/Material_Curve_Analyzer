# main.py
import json
import os
import pandas as pd
from axies_API import AxesExtractor
from curve_API import CurveAnalysisAPI

def main(image_path):
    """
    主函数：处理图像，提取坐标轴信息和曲线数据，并进行坐标转换
    
    参数:
        image_path: 图像文件路径
    """
    print("=" * 60)
    print("开始处理图像:", image_path)
    print("=" * 60)
    
    # 验证图像文件存在
    if not os.path.exists(image_path):
        print(f"错误: 图像文件不存在: {image_path}")
        return
    
    # 步骤1: 提取坐标轴信息
    print("\n步骤1: 提取坐标轴信息...")
    axes_extractor = AxesExtractor(debug=False)
    axes_info = axes_extractor.extract_axes_info(image_path)
    
    if not axes_info['success']:
        print(f"坐标轴提取失败: {axes_info['message']}")
        return
    
    # 获取格式化后的坐标轴参数
    axes_params = axes_extractor.get_formatted_results(axes_info)
    
    # 提取关键参数
    x_left = axes_params['x_axis']['limits']['left']
    x_right = axes_params['x_axis']['limits']['right']
    x_title = axes_params['x_axis']['title']
    x_pixels_per_value = axes_params['x_axis']['scale']['pixels_per_value']
    
    y_bottom = axes_params['y_axis']['limits']['bottom']
    y_top = axes_params['y_axis']['limits']['top']
    y_title = axes_params['y_axis']['title']
    y_pixels_per_value = axes_params['y_axis']['scale']['pixels_per_value']
    
    print("坐标轴信息提取完成:")
    print(f"  X轴: 左极限={x_left:.2f}, 右极限={x_right:.2f}, 标题='{x_title}', 刻度={x_pixels_per_value:.6f} 单位/像素")
    print(f"  Y轴: 下极限={y_bottom:.2f}, 上极限={y_top:.2f}, 标题='{y_title}', 刻度={y_pixels_per_value:.6f} 单位/像素")
    
    # 步骤2: 提取曲线数据
    print("\n步骤2: 提取曲线数据...")
    curve_api = CurveAnalysisAPI()
    curve_results = curve_api.analyze_image(
        image_path=image_path,
        yolo_threshold=0.3,
        curve_points=128,
        color_tolerance=40,
        visualize=False
    )
    
    if 'error' in curve_results:
        print(f"曲线提取失败: {curve_results['error']}")
        return
    
    print(f"曲线数据提取完成，找到 {curve_results['total_legends']} 条曲线")
    
    # 步骤3: 坐标转换
    print("\n步骤3: 进行坐标转换...")
    transformed_curves = []
    
    for curve in curve_results['curves']:
        legend_name = curve['legend_name']
        original_points = curve['points']
        
        print(f"  处理图例 '{legend_name}'，原始点数: {len(original_points)}")
        
        # 坐标转换
        transformed_points = []
        for point in original_points:
            # 严格按照指定的公式进行坐标转换
            # 横坐标转换: (点的横坐标 - x坐标轴左极限) × x方向每像素对应的刻度值
            new_x = (point[0] - x_left) * x_pixels_per_value
            
            # 纵坐标转换: (点的纵坐标 - y坐标轴底部极限值) × y方向每像素对应的刻度值  
            new_y = (point[1] - y_bottom) * y_pixels_per_value
            
            transformed_points.append([float(new_x), float(new_y)])
        
        transformed_curve = {
            'legend_name': legend_name,
            'original_points_count': len(original_points),
            'transformed_points': transformed_points
        }
        
        transformed_curves.append(transformed_curve)
        print(f"    坐标转换完成，转换后点数: {len(transformed_points)}")
    
    # 步骤4: 构建Excel数据
    print("\n步骤4: 构建Excel数据...")
    excel_data = []
    
    image_name = os.path.basename(image_path)
    image_id = os.path.splitext(image_name)[0]
    
    for curve in transformed_curves:
        # 将坐标点列表转换为字符串格式，便于在Excel中存储
        points_str = json.dumps(curve['transformed_points'], ensure_ascii=False)
        
        row_data = {
            'figure_name (id)': f"{image_name} ({image_id})",
            'figure_index': '',  # 空着不填
            'figure_title': '',  # 空着不填
            'x-label': x_title,
            'y-label': y_title,
            'sample': curve['legend_name'],
            'point_coordinates': points_str,
            'note': ''  # 空着不填
        }
        
        excel_data.append(row_data)
    
    # 创建DataFrame
    df = pd.DataFrame(excel_data)
    
    # 步骤5: 保存为Excel文件
    print("\n步骤5: 保存为Excel文件...")
    output_filename = f"{image_id}_result.xlsx"
    
    # 设置Excel写入器，优化格式
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Curve_Data', index=False)
        
        # 自动调整列宽
        worksheet = writer.sheets['Curve_Data']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Excel文件已保存: {output_filename}")
    
    # 步骤6: 输出结果到控制台
    print("\n" + "=" * 60)
    print("最终结果")
    print("=" * 60)
    
    print(f"图像名称: {image_name}")
    print(f"X轴标题: {x_title}")
    print(f"Y轴标题: {y_title}")
    print(f"曲线数量: {len(transformed_curves)}")
    
    # 显示Excel中的数据预览
    print("\nExcel数据预览:")
    print("-" * 100)
    print(f"{'figure_name (id)':<25} {'x-label':<15} {'y-label':<15} {'sample':<20} {'points_count':<12}")
    print("-" * 100)
    
    for row in excel_data:
        points_count = len(json.loads(row['point_coordinates']))
        print(f"{row['figure_name (id)']:<25} {row['x-label']:<15} {row['y-label']:<15} {row['sample']:<20} {points_count:<12}")
    
    return {
        'excel_data': excel_data,
        'image_name': image_name,
        'x_title': x_title,
        'y_title': y_title,
        'curves_count': len(transformed_curves)
    }

def print_detailed_coordinates(result, max_curves=3, max_points=5):
    """
    打印详细的坐标点信息
    
    参数:
        result: 主函数返回的结果字典
        max_curves: 显示的最大曲线数量
        max_points: 每条曲线显示的最大点数
    """
    print("\n" + "=" * 60)
    print("详细坐标点信息（预览）")
    print("=" * 60)
    
    excel_data = result['excel_data']
    
    for i, row in enumerate(excel_data[:max_curves]):
        points = json.loads(row['point_coordinates'])
        print(f"\n曲线 {i+1}: {row['sample']}")
        print(f"坐标点 (显示前{min(max_points, len(points))}个):")
        
        for j, point in enumerate(points[:max_points]):
            print(f"  点 {j+1:3d}: X = {point[0]:12.6f}, Y = {point[1]:12.6f}")
        
        if len(points) > max_points:
            print(f"  ... (共{len(points)}个点)")
    
    if len(excel_data) > max_curves:
        print(f"\n... (共{len(excel_data)}条曲线，完整数据请查看Excel文件)")

def create_summary_sheet(result, output_filename):
    """
    创建汇总信息工作表
    
    参数:
        result: 主函数返回的结果字典
        output_filename: 输出文件名
    """
    try:
        from openpyxl import load_workbook
        
        # 加载现有的Excel文件
        wb = load_workbook(output_filename)
        
        # 创建汇总工作表
        ws_summary = wb.create_sheet(title="Summary", index=0)
        
        # 添加汇总信息
        summary_data = [
            ["图像分析结果汇总"],
            [""],
            ["图像名称:", result['image_name']],
            ["X轴标题:", result['x_title']],
            ["Y轴标题:", result['y_title']],
            ["曲线数量:", result['curves_count']],
            [""],
            ["生成时间:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
            [""],
            ["数据说明:"],
            ["- figure_name (id): 图像名称和ID"],
            ["- figure_index: 图像索引（暂空）"],
            ["- figure_title: 图像标题（暂空）"],
            ["- x-label: X轴标题"],
            ["- y-label: Y轴标题"],
            ["- sample: 图例/样本名称"],
            ["- point_coordinates: 转换后的坐标点集合"],
            ["- note: 备注（暂空）"]
        ]
        
        # 写入汇总数据
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置标题样式
        from openpyxl.styles import Font
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        # 自动调整列宽
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # 保存修改
        wb.save(output_filename)
        print(f"已添加汇总工作表到: {output_filename}")
        
    except ImportError:
        print("警告: 未安装openpyxl，无法添加汇总工作表")
    except Exception as e:
        print(f"添加汇总工作表时出错: {e}")

if __name__ == "__main__":
    # 设置图像路径
    image_path = "1.jpg"  # 替换为您的图像路径
    
    try:
        # 执行主程序
        result = main(image_path)
        
        # 可选：打印详细坐标信息
        if result:
            print_detailed_coordinates(result, max_curves=3, max_points=5)
            
            # 添加汇总工作表
            output_filename = f"{os.path.splitext(os.path.basename(image_path))[0]}_result.xlsx"
            if os.path.exists(output_filename):
                create_summary_sheet(result, output_filename)
            
    except Exception as e:
        print(f"程序执行过程中出错: {e}")