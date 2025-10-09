# main.py
import json
import os
import pandas as pd
from axies_API import AxesExtractor
from curve_API import CurveAnalysisAPI

def safe_format(value, format_spec=".2f"):
    """
    安全格式化函数，处理None值
    
    参数:
        value: 要格式化的值
        format_spec: 格式说明符
        
    返回:
        格式化后的字符串
    """
    if value is None:
        return "None"
    try:
        return format(value, format_spec)
    except (TypeError, ValueError):
        return str(value)

def process_single_image(image_path):
    """
    处理单张图像，提取坐标轴信息和曲线数据，并进行坐标转换
    
    参数:
        image_path: 图像文件路径
        
    返回:
        处理结果字典，包含Excel数据和其他信息
    """
    print("=" * 60)
    print("开始处理图像:", image_path)
    print("=" * 60)
    
    # 验证图像文件存在
    if not os.path.exists(image_path):
        print(f"错误: 图像文件不存在: {image_path}")
        return None
    
    # 步骤1: 提取坐标轴信息
    print("\n步骤1: 提取坐标轴信息...")
    axes_extractor = AxesExtractor(debug=False)
    axes_info = axes_extractor.extract_axes_info(image_path)
    
    if not axes_info['success']:
        print(f"坐标轴提取失败: {axes_info['message']}")
        return None
    
    # 获取格式化后的坐标轴参数
    axes_params = axes_extractor.get_formatted_results(axes_info)
    
    # 提取关键参数，使用安全默认值
    x_left = axes_params['x_axis']['limits']['left'] if axes_params['x_axis']['limits']['left'] is not None else 0
    x_right = axes_params['x_axis']['limits']['right'] if axes_params['x_axis']['limits']['right'] is not None else 0
    x_title = axes_params['x_axis']['title'] if axes_params['x_axis']['title'] is not None else "未识别"
    x_pixels_per_value = axes_params['x_axis']['scale']['pixels_per_value'] if axes_params['x_axis']['scale']['pixels_per_value'] is not None else 0
    
    y_bottom = axes_params['y_axis']['limits']['bottom'] if axes_params['y_axis']['limits']['bottom'] is not None else 0
    y_top = axes_params['y_axis']['limits']['top'] if axes_params['y_axis']['limits']['top'] is not None else 0
    y_title = axes_params['y_axis']['title'] if axes_params['y_axis']['title'] is not None else "未识别"
    y_pixels_per_value = axes_params['y_axis']['scale']['pixels_per_value'] if axes_params['y_axis']['scale']['pixels_per_value'] is not None else 0
    
    print("坐标轴信息提取完成:")
    print(f"  X轴: 左极限={safe_format(x_left)}, 右极限={safe_format(x_right)}, 标题='{x_title}', 刻度={safe_format(x_pixels_per_value, '.6f')} 单位/像素")
    print(f"  Y轴: 下极限={safe_format(y_bottom)}, 上极限={safe_format(y_top)}, 标题='{y_title}', 刻度={safe_format(y_pixels_per_value, '.6f')} 单位/像素")
    
    # 检查必要的坐标轴参数是否有效
    if x_pixels_per_value == 0 or y_pixels_per_value == 0:
        print("警告: 坐标轴刻度值为0，可能影响坐标转换精度")
    
    # 步骤2: 提取曲线数据
    print("\n步骤2: 提取曲线数据...")
    curve_api = CurveAnalysisAPI()
    curve_results = curve_api.analyze_image(
        image_path=image_path,
        yolo_threshold=0.3,
        curve_points=128,
        color_tolerance=40,
        visualize=False
    )
    
    if 'error' in curve_results:
        print(f"曲线提取失败: {curve_results['error']}")
        return None
    
    print(f"曲线数据提取完成，找到 {curve_results['total_legends']} 条曲线")
    
    # 步骤3: 坐标转换
    print("\n步骤3: 进行坐标转换...")
    transformed_curves = []
    
    for curve in curve_results['curves']:
        legend_name = curve['legend_name']
        original_points = curve['points']
        
        print(f"  处理图例 '{legend_name}'，原始点数: {len(original_points)}")
        
        # 坐标转换
        transformed_points = []
        for point in original_points:
            try:
                # 严格按照指定的公式进行坐标转换
                # 横坐标转换: (点的横坐标 - x坐标轴左极限) × x方向每像素对应的刻度值
                new_x = (point[0] - x_left) * x_pixels_per_value
                
                # 纵坐标转换: (点的纵坐标 - y坐标轴底部极限值) × y方向每像素对应的刻度值  
                new_y = (point[1] - y_bottom) * y_pixels_per_value
                
                transformed_points.append([float(new_x), float(new_y)])
            except Exception as e:
                print(f"    坐标转换错误: {e}, 跳过该点")
                continue
        
        if transformed_points:
            transformed_curve = {
                'legend_name': legend_name,
                'original_points_count': len(original_points),
                'transformed_points': transformed_points
            }
            
            transformed_curves.append(transformed_curve)
            print(f"    坐标转换完成，转换后点数: {len(transformed_points)}")
        else:
            print(f"    警告: 图例 '{legend_name}' 没有成功转换的点")
    
    if not transformed_curves:
        print("警告: 没有成功转换任何曲线数据")
        return None
    
    # 步骤4: 构建Excel数据
    print("\n步骤4: 构建Excel数据...")
    excel_data = []
    
    image_name = os.path.basename(image_path)
    image_id = os.path.splitext(image_name)[0]
    
    for curve in transformed_curves:
        # 将坐标点列表转换为字符串格式，便于在Excel中存储
        points_str = json.dumps(curve['transformed_points'], ensure_ascii=False)
        
        row_data = {
            'figure_name (id)': f"{image_name} ({image_id})",
            'figure_index': '',  # 空着不填
            'figure_title': '',  # 空着不填
            'x-label': x_title,
            'y-label': y_title,
            'sample': curve['legend_name'],
            'point_coordinates': points_str,
            'note': ''  # 空着不填
        }
        
        excel_data.append(row_data)
    
    print(f"图像 {image_name} 处理完成，生成 {len(excel_data)} 行数据")
    
    return {
        'excel_data': excel_data,
        'image_name': image_name,
        'x_title': x_title,
        'y_title': y_title,
        'curves_count': len(transformed_curves),
        'image_id': image_id
    }

def process_all_images(figs_folder="figs", output_file="results.xlsx"):
    """
    处理figs文件夹中的所有图片，并将结果保存到同一个Excel文件
    
    参数:
        figs_folder: 包含图片的文件夹路径
        output_file: 输出Excel文件名
    """
    # 检查figs文件夹是否存在
    if not os.path.exists(figs_folder):
        print(f"错误: 文件夹 '{figs_folder}' 不存在")
        return
    
    # 支持的图片格式
    supported_formats = {'.jpg', '.jpeg', '.tif', '.tiff', '.png'}
    
    # 获取所有图片文件
    image_files = []
    for file in os.listdir(figs_folder):
        file_ext = os.path.splitext(file)[1].lower()
        if file_ext in supported_formats:
            image_files.append(os.path.join(figs_folder, file))
    
    # 按文件名排序
    image_files.sort()
    
    if not image_files:
        print(f"在文件夹 '{figs_folder}' 中没有找到支持的图片文件")
        print(f"支持的格式: {', '.join(supported_formats)}")
        return
    
    print(f"找到 {len(image_files)} 个图片文件:")
    for img_file in image_files:
        print(f"  - {os.path.basename(img_file)}")
    
    # 处理所有图片并收集数据
    all_excel_data = []
    processed_results = []
    failed_images = []
    
    for i, image_path in enumerate(image_files, 1):
        print(f"\n{'='*80}")
        print(f"处理第 {i}/{len(image_files)} 个图片: {os.path.basename(image_path)}")
        print(f"{'='*80}")
        
        try:
            result = process_single_image(image_path)
            if result:
                all_excel_data.extend(result['excel_data'])
                processed_results.append(result)
                print(f"✓ 成功处理: {os.path.basename(image_path)}")
            else:
                failed_images.append(os.path.basename(image_path))
                print(f"✗ 处理失败: {os.path.basename(image_path)}")
        except Exception as e:
            failed_images.append(os.path.basename(image_path))
            print(f"✗ 处理异常: {os.path.basename(image_path)} - {e}")
        
        print(f"\n完成第 {i}/{len(image_files)} 个图片的处理")
    
    if not all_excel_data:
        print("没有成功处理任何图片")
        return
    
    # 创建DataFrame并保存为Excel
    print(f"\n{'='*80}")
    print("保存所有结果到Excel文件...")
    print(f"{'='*80}")
    
    try:
        df = pd.DataFrame(all_excel_data)
        
        # 设置Excel写入器，优化格式
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='All_Curve_Data', index=0)
            
            # 自动调整列宽
            worksheet = writer.sheets['All_Curve_Data']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✓ Excel文件保存成功: {output_file}")
        
    except Exception as e:
        print(f"✗ 保存Excel文件失败: {e}")
        return
    
    # 创建汇总工作表
    create_summary_sheet(processed_results, output_file, failed_images)
    
    # 显示处理统计信息
    print_processing_summary(processed_results, failed_images)
    
    return {
        'all_excel_data': all_excel_data,
        'processed_results': processed_results,
        'failed_images': failed_images,
        'total_images': len(processed_results),
        'total_curves': len(all_excel_data)
    }

def print_processing_summary(processed_results, failed_images):
    """
    打印处理汇总信息
    
    参数:
        processed_results: 处理结果列表
        failed_images: 处理失败的图片列表
    """
    print("\n" + "=" * 80)
    print("处理汇总信息")
    print("=" * 80)
    
    print(f"成功处理的图片数量: {len(processed_results)}")
    
    if failed_images:
        print(f"处理失败的图片数量: {len(failed_images)}")
        print("失败的图片:")
        for img in failed_images:
            print(f"  - {img}")
    
    total_curves = sum(result['curves_count'] for result in processed_results)
    print(f"总曲线数量: {total_curves}")
    
    if processed_results:
        print("\n各图片处理详情:")
        print("-" * 100)
        print(f"{'图片名称':<30} {'X轴标题':<20} {'Y轴标题':<20} {'曲线数量':<10}")
        print("-" * 100)
        
        for result in processed_results:
            print(f"{result['image_name']:<30} {result['x_title']:<20} {result['y_title']:<20} {result['curves_count']:<10}")

def print_detailed_coordinates(all_excel_data, max_curves=3, max_points=5):
    """
    打印详细的坐标点信息
    
    参数:
        all_excel_data: 所有Excel数据
        max_curves: 显示的最大曲线数量
        max_points: 每条曲线显示的最大点数
    """
    print("\n" + "=" * 60)
    print("详细坐标点信息（预览）")
    print("=" * 60)
    
    for i, row in enumerate(all_excel_data[:max_curves]):
        try:
            points = json.loads(row['point_coordinates'])
            print(f"\n曲线 {i+1}: {row['sample']} (来自 {row['figure_name (id)']})")
            print(f"坐标点 (显示前{min(max_points, len(points))}个):")
            
            for j, point in enumerate(points[:max_points]):
                print(f"  点 {j+1:3d}: X = {safe_format(point[0], '12.6f')}, Y = {safe_format(point[1], '12.6f')}")
            
            if len(points) > max_points:
                print(f"  ... (共{len(points)}个点)")
        except Exception as e:
            print(f"\n曲线 {i+1} 坐标解析错误: {e}")
    
    if len(all_excel_data) > max_curves:
        print(f"\n... (共{len(all_excel_data)}条曲线，完整数据请查看Excel文件)")

def create_summary_sheet(processed_results, output_filename, failed_images):
    """
    创建汇总信息工作表
    
    参数:
        processed_results: 处理结果列表
        output_filename: 输出文件名
        failed_images: 处理失败的图片列表
    """
    try:
        from openpyxl import load_workbook
        
        # 加载现有的Excel文件
        wb = load_workbook(output_filename)
        
        # 创建汇总工作表
        ws_summary = wb.create_sheet(title="Summary", index=1)
        
        # 添加汇总信息
        summary_data = [
            ["图像分析结果汇总"],
            [""],
            ["处理时间:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["成功处理的图片数量:", len(processed_results)],
            ["处理失败的图片数量:", len(failed_images)],
            ["总曲线数量:", sum(result['curves_count'] for result in processed_results)],
            [""],
            ["成功处理的图片列表:"],
            ["图片名称", "X轴标题", "Y轴标题", "曲线数量"]
        ]
        
        # 添加每个图片的信息
        for result in processed_results:
            summary_data.append([
                result['image_name'],
                result['x_title'],
                result['y_title'],
                result['curves_count']
            ])
        
        if failed_images:
            summary_data.extend([
                [""],
                ["处理失败的图片列表:"]
            ])
            for img in failed_images:
                summary_data.append([img])
        
        summary_data.extend([
            [""],
            ["数据说明:"],
            ["- figure_name (id): 图像名称和ID"],
            ["- figure_index: 图像索引（暂空）"],
            ["- figure_title: 图像标题（暂空）"],
            ["- x-label: X轴标题"],
            ["- y-label: Y轴标题"],
            ["- sample: 图例/样本名称"],
            ["- point_coordinates: 转换后的坐标点集合"],
            ["- note: 备注（暂空）"]
        ])
        
        # 写入汇总数据
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置标题样式
        from openpyxl.styles import Font
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        # 自动调整列宽
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # 保存修改
        wb.save(output_filename)
        print(f"✓ 已添加汇总工作表到: {output_filename}")
        
    except ImportError:
        print("警告: 未安装openpyxl，无法添加汇总工作表")
    except Exception as e:
        print(f"添加汇总工作表时出错: {e}")

if __name__ == "__main__":
    try:
        # 处理figs文件夹中的所有图片
        result = process_all_images(figs_folder="figs", output_file="results.xlsx")
        
        # 可选：打印详细坐标信息
        if result:
            print_detailed_coordinates(result['all_excel_data'], max_curves=3, max_points=5)
            
    except Exception as e:
        print(f"程序执行过程中出错: {e}")
        import traceback
        traceback.print_exc()